import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, Form, UploadFile, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi import Request

from PIL import Image
import csv
import json
from openpyxl import load_workbook
import filetype

app = FastAPI(title="Universal File Converter")

BASE_DIR = Path(__file__).resolve().parent.parent
STATIC_DIR = BASE_DIR / "static"
TEMPLATES_DIR = BASE_DIR / "templates"

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


def run_cmd(cmd: list[str]) -> None:
    try:
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
    except subprocess.CalledProcessError as e:
        raise HTTPException(status_code=500, detail=f"Command failed: {' '.join(cmd)}\n{e.stderr.decode(errors='ignore')}")


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/convert")
async def convert(
    file: UploadFile = File(...),
    conversion: str = Form(...),
    target_format: Optional[str] = Form(None),
):
    src_suffix = Path(file.filename).suffix.lower()
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        src_path = tmpdir_path / f"input{src_suffix or ''}"
        with open(src_path, "wb") as f:
            f.write(await file.read())

        # Detect kind if needed
        kind = filetype.guess(str(src_path))
        mime = kind.mime if kind else None

        if conversion == "pdf_to_images":
            # Use pdftoppm (poppler-utils) to convert to PNG per page
            out_base = tmpdir_path / "page"
            cmd = ["pdftoppm", "-png", str(src_path), str(out_base)]
            run_cmd(cmd)
            # Zip all PNGs
            zip_path = tmpdir_path / "images.zip"
            shutil.make_archive(str(zip_path.with_suffix('')), 'zip', tmpdir_path, "")
            return FileResponse(path=str(zip_path), filename="pdf_images.zip")

        if conversion == "image_to_image":
            if not target_format:
                raise HTTPException(400, "target_format is required")
            tgt = target_format.lower().lstrip('.')
            out_path = tmpdir_path / f"output.{tgt}"
            # Prefer ImageMagick's convert if available, else Pillow
            if shutil.which("magick") or shutil.which("convert"):
                convert_bin = "magick" if shutil.which("magick") else "convert"
                cmd = [convert_bin, str(src_path), str(out_path)]
                run_cmd(cmd)
            else:
                im = Image.open(src_path)
                if im.mode in ("RGBA", "P") and tgt in {"jpg", "jpeg"}:
                    im = im.convert("RGB")
                im.save(out_path)
            return FileResponse(path=str(out_path), filename=out_path.name)

        if conversion == "image_to_pdf":
            out_path = tmpdir_path / "output.pdf"
            if shutil.which("magick") or shutil.which("convert"):
                convert_bin = "magick" if shutil.which("magick") else "convert"
                cmd = [convert_bin, str(src_path), str(out_path)]
                run_cmd(cmd)
            else:
                im = Image.open(src_path)
                if im.mode in ("RGBA", "P"):
                    im = im.convert("RGB")
                im.save(out_path, "PDF")
            return FileResponse(path=str(out_path), filename="output.pdf")

        if conversion == "doc_to_pdf":
            # Use LibreOffice headless
            out_dir = tmpdir_path
            cmd = [
                "libreoffice", "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(src_path)
            ]
            run_cmd(cmd)
            out_path = out_dir / (src_path.stem + ".pdf")
            if not out_path.exists():
                raise HTTPException(500, "Conversion failed: output not found")
            return FileResponse(path=str(out_path), filename=out_path.name)

        if conversion == "md_to_pdf":
            # Use pandoc to convert markdown or txt to pdf (requires LaTeX for best quality but basic PDF works via wkhtmltopdf or pandoc + wkhtmltopdf/WeasyPrint not installed). We'll rely on pandoc + wkhtmltopdf if available; else to HTML.
            out_path = tmpdir_path / "output.pdf"
            cmd = ["pandoc", str(src_path), "-o", str(out_path)]
            run_cmd(cmd)
            return FileResponse(path=str(out_path), filename="output.pdf")

        if conversion == "video_convert":
            if not target_format:
                raise HTTPException(400, "target_format is required")
            tgt = target_format.lower().lstrip('.')
            out_path = tmpdir_path / f"output.{tgt}"
            cmd = ["ffmpeg", "-y", "-i", str(src_path), str(out_path)]
            run_cmd(cmd)
            return FileResponse(path=str(out_path), filename=out_path.name)

        if conversion == "audio_convert":
            if not target_format:
                raise HTTPException(400, "target_format is required")
            tgt = target_format.lower().lstrip('.')
            out_path = tmpdir_path / f"output.{tgt}"
            # Require ffmpeg for audio conversions
            if not shutil.which("ffmpeg"):
                raise HTTPException(500, "ffmpeg is required for audio conversion")
            cmd = ["ffmpeg", "-y", "-i", str(src_path), str(out_path)]
            run_cmd(cmd)
            return FileResponse(path=str(out_path), filename=out_path.name)

        if conversion == "csv_to_json":
            out_path = tmpdir_path / "output.json"
            with open(src_path, newline='', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                rows = list(reader)
            with open(out_path, 'w', encoding='utf-8') as jf:
                json.dump(rows, jf, ensure_ascii=False)
            return FileResponse(path=str(out_path), filename=out_path.name)

        if conversion == "excel_to_json":
            out_path = tmpdir_path / "output.json"
            wb = load_workbook(src_path, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter)
            records = [dict(zip(headers, r)) for r in rows_iter]
            with open(out_path, 'w', encoding='utf-8') as jf:
                json.dump(records, jf, ensure_ascii=False)
            return FileResponse(path=str(out_path), filename=out_path.name)

        if conversion == "json_to_csv":
            out_path = tmpdir_path / "output.csv"
            with open(src_path, 'r', encoding='utf-8') as jf:
                data = json.load(jf)
            if not isinstance(data, list):
                raise HTTPException(400, "JSON must be an array of objects")
            if len(data) == 0:
                # create empty csv
                with open(out_path, 'w', newline='', encoding='utf-8') as cf:
                    pass
            else:
                fieldnames = sorted({k for item in data for k in item.keys()})
                with open(out_path, 'w', newline='', encoding='utf-8') as cf:
                    writer = csv.DictWriter(cf, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(data)
            return FileResponse(path=str(out_path), filename=out_path.name)

        if conversion == "archive_extract":
            # Extract to zip and return zip
            extract_dir = tmpdir_path / "extracted"
            extract_dir.mkdir(parents=True, exist_ok=True)
            # Try with 7z if available, else use shutil for zip, tar
            if shutil.which("7z"):
                cmd = ["7z", "x", str(src_path), f"-o{extract_dir}"]
                run_cmd(cmd)
            else:
                try:
                    shutil.unpack_archive(str(src_path), extract_dir)
                except Exception as e:
                    raise HTTPException(400, f"Unsupported archive or failed to extract: {e}")
            zip_path = tmpdir_path / "extracted.zip"
            shutil.make_archive(str(zip_path.with_suffix('')), 'zip', extract_dir)
            return FileResponse(path=str(zip_path), filename="extracted.zip")

        if conversion == "archive_create":
            # If the input is a directory in an archive format, we can repackage
            # For simplicity, if input is a text file listing files (one per line), zip them
            out_path = tmpdir_path / "archive.zip"
            with tempfile.TemporaryDirectory() as files_dir:
                # If it's a zip already, just return it
                if src_suffix in {".zip", ".tar", ".tar.gz", ".tgz"}:
                    return FileResponse(path=str(src_path), filename=Path(file.filename).name)
                # Otherwise, create a zip with just this file
                shutil.copy(str(src_path), str(Path(files_dir) / Path(file.filename).name))
                shutil.make_archive(str(out_path.with_suffix('')), 'zip', files_dir)
            return FileResponse(path=str(out_path), filename=out_path.name)

        raise HTTPException(400, f"Unknown conversion: {conversion}")